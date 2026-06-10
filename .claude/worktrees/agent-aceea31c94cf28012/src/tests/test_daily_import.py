"""Unit tests for DailyImportJob stages (validate, transform, report)."""
import base64
import io
import tempfile
from datetime import date
from pathlib import Path
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from openpyxl import load_workbook

from src.reporting.jobs.daily_import import DailyImportJob, _read_csv, ERROR_RATE_THRESHOLD


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_clean_records(n: int) -> list[dict]:
    return [
        {
            "id": str(i),
            "product_code": "PROD001",
            "date": "2024-01-15",
            "quantity": "10",
            "unit_price": "9.99",
        }
        for i in range(n)
    ]


def _make_quality(error_rate: float = 0.0, total: int = 100) -> dict:
    errors = int(total * error_rate)
    return {
        "total_records": total,
        "valid_count": total - errors,
        "error_count": errors,
        "file_count": 2,
        "null_count": 0,
        "duplicate_count": 0,
        "negative_count": 0,
        "invalid_code_count": errors,
        "error_rate": error_rate,
        "passed": error_rate <= ERROR_RATE_THRESHOLD,
    }


# ---------------------------------------------------------------------------
# Validate stage
# ---------------------------------------------------------------------------


class TestValidateStage:
    @pytest.fixture
    def job(self):
        return DailyImportJob()

    async def test_clean_records_all_valid(self, job):
        records = _make_clean_records(10)
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["valid_count"] == 10
        assert result["quality"]["error_count"] == 0
        assert result["quality"]["passed"] is True

    async def test_null_product_code_flagged(self, job):
        records = _make_clean_records(10)
        records[0]["product_code"] = None
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["null_count"] == 1
        assert len(result["error_rows"]) == 1
        assert "null_product_code" in result["error_rows"][0]["_errors"]

    async def test_empty_product_code_flagged(self, job):
        records = _make_clean_records(5)
        records[0]["product_code"] = ""
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["null_count"] == 1

    async def test_invalid_product_code_flagged(self, job):
        records = _make_clean_records(5)
        records[0]["product_code"] = "INVALID"
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["invalid_code_count"] == 1
        assert "invalid_product_code" in result["error_rows"][0]["_errors"]

    async def test_duplicate_id_flagged(self, job):
        records = _make_clean_records(5)
        records[1]["id"] = records[0]["id"]  # duplicate
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["duplicate_count"] == 1

    async def test_negative_quantity_flagged(self, job):
        records = _make_clean_records(5)
        records[0]["quantity"] = "-5"
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["negative_count"] == 1
        assert "negative_quantity" in result["error_rows"][0]["_errors"]

    async def test_error_rate_exceeds_threshold_raises(self, job):
        # 20% errors — over 10% threshold
        records = _make_clean_records(10)
        for i in range(2):
            records[i]["product_code"] = "INVALID"
        state = {"records": records, "file_count": 1}
        with pytest.raises(ValueError, match="Data quality failed"):
            await job.validate({}, state)

    async def test_error_rate_at_threshold_passes(self, job):
        # Exactly 10% errors → passes (<=)
        records = _make_clean_records(10)
        records[0]["product_code"] = "INVALID"  # 1/10 = 10%
        state = {"records": records, "file_count": 1}
        result = await job.validate({}, state)
        assert result["quality"]["passed"] is True

    async def test_empty_records_passes(self, job):
        state = {"records": [], "file_count": 0}
        result = await job.validate({}, state)
        assert result["quality"]["total_records"] == 0
        assert result["quality"]["passed"] is True


# ---------------------------------------------------------------------------
# Transform stage
# ---------------------------------------------------------------------------


class TestTransformStage:
    @pytest.fixture
    def job(self):
        return DailyImportJob()

    async def test_transforms_all_records(self, job):
        records = _make_clean_records(5)
        state = {"valid_records": records}
        result = await job.transform({}, state)
        assert len(result["transformed"]) == 5

    async def test_quantity_cast_to_int(self, job):
        records = [{"id": "1", "product_code": "PROD001", "date": "2024-01-01", "quantity": "42", "unit_price": "5.0"}]
        state = {"valid_records": records}
        result = await job.transform({}, state)
        assert result["transformed"][0]["quantity"] == 42
        assert isinstance(result["transformed"][0]["quantity"], int)

    async def test_unit_price_cast_to_float(self, job):
        records = [{"id": "1", "product_code": "PROD001", "date": "2024-01-01", "quantity": "1", "unit_price": "9.99"}]
        state = {"valid_records": records}
        result = await job.transform({}, state)
        assert result["transformed"][0]["unit_price"] == pytest.approx(9.99)

    async def test_date_preserved_as_string(self, job):
        records = [{"id": "1", "product_code": "PROD001", "date": "2024-06-15", "quantity": "1", "unit_price": "1.0"}]
        state = {"valid_records": records}
        result = await job.transform({}, state)
        assert result["transformed"][0]["date"] == "2024-06-15"

    async def test_empty_records(self, job):
        state = {"valid_records": []}
        result = await job.transform({}, state)
        assert result["transformed"] == []


# ---------------------------------------------------------------------------
# Report stage
# ---------------------------------------------------------------------------


class TestReportStage:
    @pytest.fixture
    def job(self):
        return DailyImportJob()

    async def test_returns_base64_encoded_bytes(self, job):
        quality = _make_quality(0.0, 10)
        state = {"quality": quality, "valid_records": _make_clean_records(10), "error_rows": []}
        result = await job.report({}, state)
        b64 = result["xlsx_bytes_b64"]
        xlsx_bytes = base64.b64decode(b64)
        assert len(xlsx_bytes) > 0

    async def test_returns_valid_xlsx(self, job):
        quality = _make_quality(0.0, 10)
        state = {"quality": quality, "valid_records": _make_clean_records(10), "error_rows": []}
        result = await job.report({}, state)
        xlsx_bytes = base64.b64decode(result["xlsx_bytes_b64"])
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        assert "Quality Summary" in wb.sheetnames

    async def test_summary_keys_present(self, job):
        quality = _make_quality(0.0, 10)
        state = {"quality": quality, "valid_records": _make_clean_records(10), "error_rows": []}
        result = await job.report({}, state)
        for key in ("total_records", "valid_count", "error_count", "error_rate", "passed"):
            assert key in result["summary"]


# ---------------------------------------------------------------------------
# _read_csv helper
# ---------------------------------------------------------------------------


class TestReadCsv:
    async def test_reads_csv_to_list_of_dicts(self, tmp_path):
        p = tmp_path / "test.csv"
        p.write_text("id,product_code,date,quantity,unit_price\n1,PROD001,2024-01-01,10,9.99\n")
        records = await _read_csv(str(p))
        assert len(records) == 1
        assert records[0]["id"] == "1"
        assert records[0]["product_code"] == "PROD001"

    async def test_reads_multiple_rows(self, tmp_path):
        p = tmp_path / "test.csv"
        p.write_text("id,product_code\n1,PROD001\n2,PROD002\n3,PROD003\n")
        records = await _read_csv(str(p))
        assert len(records) == 3

    async def test_returns_list_of_dicts(self, tmp_path):
        p = tmp_path / "test.csv"
        p.write_text("id,val\n1,a\n")
        records = await _read_csv(str(p))
        assert isinstance(records[0], dict)


# ---------------------------------------------------------------------------
# End-to-end pipeline (no DB) — discover is skipped
# ---------------------------------------------------------------------------


class TestDailyImportPipelineNoDb:
    async def test_full_run_produces_artifact(self, tmp_path):
        """Run all stages (skipping discover & archive which need DB) via direct state wiring."""
        job = DailyImportJob()

        # Manually build state as if discover+ingest already ran
        from src.reporting.generators.csv_data import generate_csv_files

        csv_paths = generate_csv_files(
            seed=7, n_files=2, rows_per_file=50, inject_errors=2, output_dir=tmp_path / "csv"
        )

        # ingest
        ingest_state: dict = {}
        ingest_result = await job.ingest({}, {"files": [str(p) for p in csv_paths]})
        ingest_state.update(ingest_result)
        ingest_state["file_count"] = len(csv_paths)

        # validate
        validate_result = await job.validate({}, ingest_state)
        combined = {**ingest_state, **validate_result}

        # transform
        transform_result = await job.transform({}, combined)
        combined.update(transform_result)

        # report
        report_result = await job.report({}, combined)
        assert "xlsx_bytes_b64" in report_result
        assert report_result["summary"]["passed"] is True
