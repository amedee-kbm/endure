"""Unit tests for src.reporting.generators.excel."""
import io

import pytest
from openpyxl import load_workbook

from src.reporting.generators.excel import render_excel


SAMPLE_QUALITY = {
    "total_records": 100,
    "valid_count": 95,
    "error_count": 5,
    "file_count": 2,
    "null_count": 1,
    "duplicate_count": 2,
    "negative_count": 1,
    "invalid_code_count": 1,
    "error_rate": 0.05,
    "passed": True,
}

SAMPLE_VALID = [{"id": "1", "product_code": "PROD001", "date": "2024-01-01", "quantity": 10, "unit_price": 9.99}]

SAMPLE_ERRORS = [{"id": "2", "product_code": "INVALID", "date": "2024-01-01", "quantity": 5, "unit_price": 1.0, "_errors": "invalid_product_code"}]


class TestRenderExcel:
    def _load_wb(self, result: bytes):
        return load_workbook(io.BytesIO(result))

    def test_returns_bytes(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_two_sheets(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        assert len(wb.sheetnames) == 2

    def test_sheet_names(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        assert wb.sheetnames[0] == "Quality Summary"
        assert wb.sheetnames[1] == "Exception Detail"

    def test_quality_summary_has_metric_column(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        ws = wb["Quality Summary"]
        # First row header
        assert ws.cell(1, 1).value == "Metric"
        assert ws.cell(1, 2).value == "Value"

    def test_quality_summary_total_records(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        ws = wb["Quality Summary"]
        # row 2 = Total Records
        assert ws.cell(2, 1).value == "Total Records"
        assert ws.cell(2, 2).value == 100

    def test_quality_summary_pass_status(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        ws = wb["Quality Summary"]
        # Last row = Pass
        pass_row = [ws.cell(r, 1).value for r in range(1, ws.max_row + 1)].index("Pass") + 1
        assert ws.cell(pass_row, 2).value == "True"

    def test_exception_detail_has_error_rows(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        ws = wb["Exception Detail"]
        # Header row + 1 data row
        assert ws.max_row >= 2

    def test_no_errors_shows_no_errors_message(self):
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, [])
        wb = self._load_wb(result)
        ws = wb["Exception Detail"]
        assert ws.cell(1, 1).value == "No errors"

    def test_valid_xlsx_format(self):
        """Bytes should be a valid xlsx that openpyxl can load without errors."""
        result = render_excel(SAMPLE_QUALITY, SAMPLE_VALID, SAMPLE_ERRORS)
        wb = self._load_wb(result)
        assert wb is not None
